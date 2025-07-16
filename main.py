# -*- encoding: utf-8 -*-
'''
@FileName : rsas_xml_report_export_tools.py
@Description : transform zip of RSAS xml report to xlsx report
@Date : 2025/07/17 01:20:36
@Author : Kevin
@version : 2.0
@Contact : chlitsec@163.com
'''
import logging
from datetime import date
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from sys import argv
from zipfile import ZipFile

# log configuration
logger = logging.getLogger()
logger.setLevel(logging.INFO)
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
log_file = "export_tools.log"
fh = logging.FileHandler(log_file, encoding='utf-8')
fh.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
ch.setFormatter(formatter)
fh.setFormatter(formatter)
logger.addHandler(ch)
logger.addHandler(fh)


def GetRiskLevel(risk_points):
    return (
        "低危" if risk_points >= 0 and risk_points < 4 else
        "中危" if risk_points >= 4 and risk_points < 7 else
        "高危" if risk_points >= 7 and risk_points <= 10 else
        None
    ) if risk_points is not None else None


def parse_xml_to_xlsx(xml_file):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
    except FileNotFoundError:
        logging.error(f"Error: XML file not found at {xml_file}")
        return
    except ET.ParseError as e:
        logging.error(f"Error parsing XML file: {e}")
        return
    data_list = []
    serial_number = 1
    targets = root.findall(".//data/report/targets/target")
    fileName = root.find(".//data/report/task/name").text
    for target in targets:
        ip_element = target.find("ip")
        ip_address = ip_element.text if ip_element is not None else "N/A"
        vuln_detail_elements = target.find("vuln_detail")
        if vuln_detail_elements is not None:
            for detail_vuln in vuln_detail_elements.findall("vuln"):
                vul_id = detail_vuln.find("vul_id").text if detail_vuln.find("vul_id") is not None else "N/A"
                name = detail_vuln.find("name").text if detail_vuln.find("name") is not None else "N/A"
                risk_points = detail_vuln.find("risk_points").text if detail_vuln.find("risk_points") is not None else "N/A"
                solution = detail_vuln.find("solution").text if detail_vuln.find("solution") is not None else "N/A"
                description = detail_vuln.find("description").text if detail_vuln.find("description") is not None else "N/A"
                vuln_details_lookup[vul_id] = {
                    "name": name,
                    "level": GetRiskLevel(float(risk_points)), 
                    "description": description,
                    "solution": solution
                }
        vuln_scanned_elements = target.find("vuln_scanned")
        if vuln_scanned_elements is not None:
            for scanned_vuln in vuln_scanned_elements.findall("vuln"):
                port = scanned_vuln.find("port").text if scanned_vuln.find("port") is not None and scanned_vuln.find("port").text != '0'else ""
                vul_id = scanned_vuln.find("vul_id").text if scanned_vuln.find("vul_id") is not None else "N/A"

                # Get the corresponding detail information
                detail_info = vuln_details_lookup.get(vul_id, {})  # Get dict or empty dict if not found
                row = {
                    "序号": serial_number,
                    "风险名称": detail_info.get("name", "N/A"),
                    "风险等级": detail_info.get("level", "N/A"),
                    "风险描述": detail_info.get("description", "N/A"),
                    "加固建议": detail_info.get("solution", "N/A"),
                    "风险来源": ip_address,
                    "端口": port,
                    "扫描设备": f"绿盟RSAS (漏洞库版本{root.find(".//data/report/sysvul_version").text})",
                }
                if row["风险等级"] != "低危":
                    data_list.append(row)
                    serial_number += 1
                else:
                    pass
        else:
            pass
    fieldnames = ["序号", "风险名称", "风险等级", "风险描述", "加固建议", "风险来源", "端口", "扫描设备"]
    wb = Workbook()
    ws = wb.active
    ws.title = "漏洞清单"  # 自定义工作表名称
    # --- set workbook tittle style ---
    # set font
    header_font = Font(
        bold=True,    # 加粗
    )
    # set fill
    header_fill = PatternFill(
        start_color='4F81BD',  # 背景颜色 (深蓝色)
        end_color='4F81BD',
        fill_type='solid'
    )
    # set aligment
    header_alignment = Alignment(
        horizontal='center',  # 水平居中
        vertical='center'    # 垂直居中
    )
    # set border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header_name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data_dict in enumerate(data_list, 2):
        for col_idx, field_name in enumerate(fieldnames, 1):
            # 使用 .get() 方法安全地获取字典值，如果键不存在则默认为 None 或空字符串
            cell_value = row_data_dict.get(field_name, '')
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    try:
        wb.save(fileName+'_'+str(date.today())+'.xlsx')
        logging.info(f"Excel文件 '{fileName+'_'+str(date.today())}' 已转换完成。")
        # safe_rename_file(xml_file, fileName)
    except Exception as e:
        logging.error(f"保存Excel文件时出错: {e}")

if __name__ == '__main__':
    if len(argv) == 2 and argv[1].endswith('.zip'):
        with ZipFile(argv[1], 'r') as zip_ref:
            logging.info(f"ZIP文件 '{argv[1]}' 正在进行转换")
            for file_info in zip_ref.filelist:
                if file_info.filename.endswith('.xml'):
                    with zip_ref.open(file_info.filename) as xml_file:
                        try:
                            parse_xml_to_xlsx(xml_file)
                        except ET.ParseError as e:
                            logging.error(f"解析失败: {file_info.filename} - {e}")
                    pass
        logging.info(f"ZIP文件 '{argv[1]}' 转换完成")
    else:
        print(f"Usage: python {argv[0]} <zip file>")
        print("将RSAS导出的XML压缩包转换为xlsx格式.")
